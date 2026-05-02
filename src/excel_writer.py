from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import OUTPUT_DIR as DEFAULT_OUTPUT_DIR

OUTPUT_COLUMNS = [
    "물품식별번호",
    "검색결과순번",
    "동일번호결과수",
    "상품명",
    "업체명",
    "계약번호",
    "계약변경차수",
    "계약단가",
    "계약기간",
    "MAS여부",
    "계약유형",
    "공급지역",
    "인도조건",
    "쇼핑몰등록일자",
    "업체사업자등록번호",
    "상세페이지링크",
    "구성",
    "옵션/기타",
    "수집방식",
    "처리상태",
    "오류내용"
]

# 기본 출력 위치
CURRENT_OUTPUT_DIR = DEFAULT_OUTPUT_DIR
OUTPUT_EXCEL_PATH = CURRENT_OUTPUT_DIR / "result.xlsx"
FAILED_EXCEL_PATH = CURRENT_OUTPUT_DIR / "failed_result.xlsx"


def set_output_dir(output_dir):
    """
    GUI에서 선택한 결과 폴더로 저장 위치를 변경한다.
    """

    global CURRENT_OUTPUT_DIR, OUTPUT_EXCEL_PATH, FAILED_EXCEL_PATH

    CURRENT_OUTPUT_DIR = Path(output_dir)
    CURRENT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    OUTPUT_EXCEL_PATH = CURRENT_OUTPUT_DIR / "result.xlsx"
    FAILED_EXCEL_PATH = CURRENT_OUTPUT_DIR / "failed_result.xlsx"


def get_output_paths():
    """
    현재 결과 파일 경로를 반환한다.
    """

    return OUTPUT_EXCEL_PATH, FAILED_EXCEL_PATH


def normalize_rows(rows):
    """
    rows 데이터를 DataFrame으로 바꾸기 전에 컬럼 누락을 방지한다.
    """

    normalized = []

    for row in rows:
        new_row = {}

        for column in OUTPUT_COLUMNS:
            new_row[column] = row.get(column, "")

        normalized.append(new_row)

    return normalized


def apply_excel_style(file_path):
    """
    저장된 엑셀 파일에 보기 좋은 스타일을 적용한다.
    """

    wb = None

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        header_font = Font(
            bold=True
        )

        normal_alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

        link_font = Font(
            color="0000FF",
            underline="single"
        )

        # 헤더 스타일
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        # 컬럼 인덱스 찾기
        header_map = {}

        for col_index, cell in enumerate(ws[1], start=1):
            header_map[cell.value] = col_index

        link_col_index = header_map.get("상세페이지링크")

        # 본문 스타일
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = normal_alignment

            if link_col_index:
                link_cell = row[link_col_index - 1]
                link_value = str(link_cell.value).strip() if link_cell.value else ""

                if link_value.startswith("http://") or link_value.startswith("https://"):
                    link_cell.hyperlink = link_value
                    link_cell.font = link_font

        # 컬럼 너비 자동 조정
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            header_value = column_cells[0].value

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                value = str(value)
                longest_line = max(value.splitlines(), key=len, default="")
                max_length = max(max_length, len(longest_line))

            if header_value in ["구성", "옵션/기타", "오류내용"]:
                width = min(max(max_length + 2, 20), 60)
            elif header_value == "상세페이지링크":
                width = min(max(max_length + 2, 25), 70)
            else:
                width = min(max(max_length + 2, 12), 35)

            ws.column_dimensions[column_letter].width = width

        for row_index in range(2, ws.max_row + 1):
            ws.row_dimensions[row_index].height = 45

        ws.freeze_panes = "A2"

        wb.save(file_path)

    finally:
        if wb is not None:
            wb.close()


def save_dataframe_to_excel(df, file_path, sheet_name):
    """
    DataFrame을 엑셀로 저장하고 스타일을 적용한다.
    """

    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name
        )

    apply_excel_style(file_path)


def save_failed_rows(rows):
    """
    실패한 행만 failed_result.xlsx로 따로 저장한다.
    """

    failed_rows = []

    for row in rows:
        status = str(row.get("처리상태", ""))

        if "실패" in status or row.get("오류내용"):
            failed_rows.append(row)

    if not failed_rows:
        return

    df_failed = pd.DataFrame(normalize_rows(failed_rows))
    df_failed = df_failed[OUTPUT_COLUMNS]

    save_dataframe_to_excel(
        df_failed,
        FAILED_EXCEL_PATH,
        "실패목록"
    )

    print(f"실패 목록 저장 완료: {FAILED_EXCEL_PATH}")


def save_to_excel(rows):
    """
    수집한 데이터를 result.xlsx 파일로 저장한다.
    실패한 항목이 있으면 failed_result.xlsx에도 따로 저장한다.
    """

    if not rows:
        print("저장할 데이터가 없습니다.")
        return

    try:
        normalized_rows = normalize_rows(rows)

        df = pd.DataFrame(normalized_rows)
        df = df[OUTPUT_COLUMNS]

        save_dataframe_to_excel(
            df,
            OUTPUT_EXCEL_PATH,
            "수집결과"
        )

        print(f"엑셀 저장 완료: {OUTPUT_EXCEL_PATH}")

        save_failed_rows(normalized_rows)

    except PermissionError as error:
        print(f"엑셀 저장 실패: {error}")
        raise PermissionError(
            f"결과 엑셀 파일을 저장할 수 없습니다.\n"
            f"파일이 열려 있거나 OneDrive/탐색기 미리보기에서 사용 중일 수 있습니다.\n"
            f"대상 파일: {OUTPUT_EXCEL_PATH}\n"
            f"상세 오류: {error}"
        )